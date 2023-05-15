import openpyxl
import datetime
import re
import tkinter
from tkinter import *
from tkinter import filedialog, scrolledtext
from sql import *

def loaded():
    global report
    report = filedialog.askopenfilename()
    return report

time = datetime.date.today()

def bal_ids():
    # спрашиваем логин и пароль

    login_input = "{}".format(login.get())
    password_input = "{}".format(password.get())

    # коннектимся
    dsn_tns = cx_Oracle.makedsn('#####', '####', '####')
    conn = cx_Oracle.connect(user=login_input, password=password_input,
                             dsn=dsn_tns)
    c = conn.cursor()

    wb2 = openpyxl.reader.excel.load_workbook(report)

    wb2.remove(wb2['Transaction Details'])
    wb2.remove(wb2['Legend'])

    emails = ''
    converted_emails =''

    # Тащим нужный лист из таблицы
    ws2 = wb2.active
    ws2.delete_cols(3, 2)

    # вытаскиваем адреса

    i2 = 2
    while (ws2['B' + str(i2)].value != None):
        emails += ("'" + str(ws2['A' + str(i2)].value) + "',")
        i2 = i2 + 1
    emails = emails[0:-1]
    print(emails)

    request_bal_id = f"""###"""

    c.execute(request_bal_id)

    for row in c:
        peprow = str(row)[2:-3]
        converted_emails += '\n '
        converted_emails += peprow
    conn.close()

    print(converted_emails)
    txt.insert(INSERT, converted_emails)

    #return

def start():
    #спрашиваем логин и пароль

    login_input = "{}".format(login.get())
    password_input = "{}".format(password.get())

    login_input = "########"
    password_input = "#############"

    #коннектимся
    dsn_tns = cx_Oracle.makedsn('####', '####', '####')
    conn = cx_Oracle.connect(user=login_input, password=password_input,
                             dsn=dsn_tns)
    c = conn.cursor()

    wb = openpyxl.reader.excel.load_workbook(report)
    wb.remove(wb['Legend'])
    wb.remove(wb['Summary'])

    ticket = "{}".format(input_ticket_number.get())

    pay_id_with_A = ''
    converted_pay_id = ''


    # Тащим нужный лист
    ws = wb.active

    print(ws)

    ws.delete_cols(2,10)
    ws.delete_cols(3,12)

    ws['D1'] = 'Date Report'
    ws['E1'] = 'Ticket'

    # ищем необходимую запись с 'A'
    i = 2
    while (ws['B' + str(i)].value != None):
        if (ws['B' + str(i)].value)[0] == 'A':
            if (ws['B' + str(i + 1)].value != None):
                pay_id_with_A += ("'" + str(ws['C' + str(i)].value) + "',")
        ws['D' + str(i)].value = time
        ws['E' + str(i)].value = ticket
        i = i + 1
    pay_id_with_A = pay_id_with_A[0:-1]

    print(len(pay_id_with_A))


    request_pay_id = f"""
    ###"""

    c.execute(request_pay_id)

    for row in c:
        peprow = str(row)[1:-2]
        converted_pay_id += ' '
        converted_pay_id += peprow
    conn.close()

    splitted_str = converted_pay_id.split()
    print(splitted_str)
    print(len(splitted_str))
    print(converted_pay_id)
    print(len(converted_pay_id))


    i1 = 2
    while (ws['B' + str(i1)].value != None):
        p = 0
        if (ws['B' + str(i1)].value)[0] == 'A':
            ws['B' + str(i1)].value = splitted_str[p]
            p += 1
        i1 = i1 + 1


    # сохраняем файл
    f = filedialog.asksaveasfile(mode = 'w', defaultextension = '.xlsx', filetypes=[
        ('.xlsx', '.xlsx')
    ])
    saved_file = list(re.findall(r"'([^']+)'", str(f)))[0]
    wb.save(saved_file)

    print(pay_id_with_A)
    return


window = tkinter.Tk()
window.title("Project")
# window.geometry('600x400')
window.geometry('1000x400')


lbl1 = Label(window, text="Выбери файл", font=("Trebuchet MS", 12))
lbl1.place(x = 0, y = 20, width = 600)
btn1 = Button(window, text="Выбрать!", font=("Trebuchet_MS", 9), command=loaded)
btn1.place(x = 250, y = 50, width = 100)
lbl2 = Label(window, text="Введи номер задачи:", font=("Trebuchet MS", 12))
lbl2.place(x = 0, y = 80, width = 600)
input_ticket_number = Entry(window, width=15)
input_ticket_number.place(x = 250, y = 110, width = 100)
lbl3 = Label(window, text="Введи свой логин:", font=("Trebuchet MS", 12))
lbl3.place(x = 0, y = 140, width = 600)
login = Entry(window, width=15)
login.place(x = 250, y = 170, width = 100)
lbl4 = Label(window, text="Введи свой пароль:", font=("Trebuchet MS", 12))
lbl4.place(x = 0, y = 200, width = 600)
password = Entry(window, width=15)
password.place(x = 250, y = 230, width = 100)
btn5 = Button(window, text="Сформировать таблицу!", font=("Trebuchet_MS", 9), command=start)
btn5.place(x = 200, y = 280, width = 200)
btn5 = Button(window, text="Выгрузить информацию!", font=("Trebuchet_MS", 9), command=bal_ids)
btn5.place(x = 200, y = 310, width = 200)
txt = scrolledtext.ScrolledText(window, width=40, height=10)
txt.place(x = 550, y = 30, width = 400, height= 300)
window.mainloop()